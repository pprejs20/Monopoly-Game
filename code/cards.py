import openpyxl
from tile import get_cell_ref


class Card:
    def __init__(self, description):
        self.description = description

    def get_description(self):
        return self.description


class TransactionCard(Card):
    def __init__(self, payer, reciever, amount, description):
        self.payer = payer
        self.reciever = reciever
        self.amount = amount
        super().__init__(description)

    def execute(self, player, players, game):
        if self.reciever == "bank":
            player.deduct_money(self.amount)
        if self.payer == "bank":
            player.add_money(self.amount)
        if self.payer == "all":
            for i in range(len(players)-1):
                assert players.get(i).name != player.name
                p = players.get(i)
                p.deduct_money(self.amount)
                player.add_money(self.amount)
        if self.reciever == "freeparking":
            player.deduct_money(self.amount)
            game.free_parking_money += self.amount


    @classmethod
    def load_transaction_card_from_xlsx(cls, path="ExcelData/PropertyTycoonCardDataRefactored.xlsx"):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        pot_cards = []
        opp_cards = []

        for row in range(4, 17):
            row_data = []
            for col in range(1, 5):

                cell = get_cell_ref(row, col)

                row_data.append(sheet[cell].value)
            pot_cards.append(TransactionCard(row_data[0], row_data[1], row_data[2], row_data[3]))

        for row in range(29, 35):
            row_data = []
            for col in range(1, 5):

                cell = get_cell_ref(row, col)

                row_data.append(sheet[cell].value)
            opp_cards.append(TransactionCard(row_data[0], row_data[1], row_data[2], row_data[3]))

        return pot_cards, opp_cards

    def __str__(self):
        string  = "----------- Transaction Card -----------\n"
        string += "Payer: {}\n".format(self.payer)
        string += "Reciever: {}\n".format(self.reciever)
        string += "Amount: {}\n".format(self.amount)
        string += "Description: {}\n".format(self.description)
        string += "----------------------------"
        return string


class MovementCard(Card):
    def __init__(self, relative_pos, tile, pass_go, description):
        self.relative_pos = relative_pos
        self.tile = tile
        self.pass_go = pass_go
        super().__init__(description)

    def execute(self, player):
        if self.relative_pos:
            player.move_player(self.tile-1)
        else:
            if self.tile > 0:
             player.move_player_forward(self.tile) ##add new method to move forward or backward with minus
            else:
                player.move_player_backward(self.tile * -1)
        if player.pos > self.tile-1:
            if self.pass_go:
                player.add_money(200)

    @classmethod
    def load_movement_card_from_xlsx(cls, path="ExcelData/PropertyTycoonCardDataRefactored.xlsx"):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        pot_cards = []
        opp_cards = []

        for row in range(19, 22):
            row_data = []
            for col in range(1, 5):
                cell = get_cell_ref(row, col)

                row_data.append(sheet[cell].value)
            pot_cards.append(MovementCard(row_data[0], row_data[1], row_data[2], row_data[3]))

        for row in range(41, 48):
            row_data = []
            for col in range(1, 5):
                cell = get_cell_ref(row, col)

                row_data.append(sheet[cell].value)
            opp_cards.append(MovementCard(row_data[0], row_data[1], row_data[2], row_data[3]))

        return pot_cards, opp_cards

    def __str__(self):
        string = "----------- Movement Card -----------\n"
        string += "Relative position: {}\n".format(self.relative_pos)
        string += "Tile: {}\n".format(self.tile)
        string += "Pass go: {}\n".format(self.pass_go)
        string += "Description: {}\n".format(self.description)
        string += "----------------------------"
        return string


class HouseHotelCard(Card):
    def __init__(self, payer, reciever, house_amount, hotel_amount, description):
        self.payer = payer
        self.reciever = reciever
        self.house_amount = house_amount
        self.hotel_amount = hotel_amount
        super().__init__(description)

    def execute(self, player):
        houseCount = 0
        hotelCount = 0
        for prop in player.propList:
            houseCount += prop.no_of_houses
            hotelCount += prop.no_of_hotel
        player.deduct_money(houseCount*self.house_amount)
        player.deduct_money(hotelCount*self.hotel_amount)


    @classmethod
    def load_hh_card_from_xlsx(cls, path="ExcelData/PropertyTycoonCardDataRefactored.xlsx"):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        cards = []

        for row in range(37, 39):
            row_data = []
            for col in range(1, 6):
                cell = get_cell_ref(row, col)

                row_data.append(sheet[cell].value)
            cards.append(HouseHotelCard(row_data[0], row_data[1], row_data[2], row_data[3], row_data[4]))

        return cards

    def __str__(self):
        string = "----------- HH Card -----------\n"
        string += "Payer: {}\n".format(self.payer)
        string += "Reciever: {}\n".format(self.reciever)
        string += "House amount: {}\n".format(self.house_amount)
        string += "Hotel amount: {}\n".format(self.hotel_amount)
        string += "Description: {}\n".format(self.description)
        string += "----------------------------"
        return string


class JailfreeCard(Card):
    def __init__(self, description):
        super().__init__(description)

    @classmethod
    def load_jail_card_from_xlsx(cls, path="ExcelData/PropertyTycoonCardDataRefactored.xlsx"):
        workbook = openpyxl.load_workbook(path)
        sheet = workbook.active
        pot_cards = []
        opp_cards = []

        for row in range(24, 25):
            row_data = []
            for col in range(1, 2):
                cell = get_cell_ref(row, col)

                row_data.append(sheet[cell].value)
            pot_cards.append(JailfreeCard(row_data[0]))

        for row in range(50, 51):
            row_data = []
            for col in range(1, 2):
                cell = get_cell_ref(row, col)

                row_data.append(sheet[cell].value)
            opp_cards.append(JailfreeCard(row_data[0]))

        return pot_cards, opp_cards

    def __str__(self):
        string = "----------- Jail Free Card -----------\n"
        string += "Description: {}\n".format(self.description)
        string += "----------------------------"
        return string


def load_all_cards():
    potluck = []
    opportunity = []

    pot_trans, opp_trans = TransactionCard.load_transaction_card_from_xlsx()
    pot_move, opp_move = MovementCard.load_movement_card_from_xlsx()
    pot_jail, opp_jail = JailfreeCard.load_jail_card_from_xlsx()
    opp_hh = HouseHotelCard.load_hh_card_from_xlsx()

    for card in pot_trans:
        potluck.append(card)
    for card in pot_move:
        potluck.append(card)
    for card in pot_jail:
        potluck.append(card)
    # print("-------------Potluck Cards-------------\n")
    # for card in potluck:
    #     print(card)

    for card in opp_trans:
        opportunity.append(card)
    for card in opp_move:
        opportunity.append(card)
    for card in opp_hh:
        opportunity.append(card)
    for card in opp_jail:
        opportunity.append(card)

    # print("-------------Opportunity Knocks Cards-------------\n")
    # for card in opportunity:
    #     print(card)

    return potluck, opportunity


load_all_cards()



# card = JailfreeCard("jailfree")
# print(isinstance(card, Card))



# cards1, cards2 = load_all_cards()
# for c in cards1:
#     print(c)
# for c1 in cards2:
#     print(c1)

# cards1, cards2 = MovementCard.load_movement_card_from_xlsx()
# for c in cards1:
#     print(c)
#     print(type(c.tile))
# for c1 in cards2:
#     print(c1)
#     print(type(c.tile))

# cards1, cards2 = JailfreeCard.load_jail_card_from_xlsx()
# for c in cards1:
#     print(c)
# for c1 in cards2:
#     print(c1)

# cards4 = HouseHotelCard.load_hh_card_from_xlsx()
# for c in cards4:
#     print(c)









